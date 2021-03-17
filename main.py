import openpyxl
import numpy as np
import PySimpleGUI as sg
import xlsxwriter
from datetime import date
agList = {"2019-ag-6051": "FATIMA BASHIR", "2019-ag-6052": "ANIQA FAYYAZ", "2019-ag-6053": "AHTISHAM", "2019-ag-6054": "MUHAMMAD ABUBAKR", "2019-ag-6055": "JAHANZAIB BABAR", "2019-ag-6056": "MUHAMMAD SHOAIB AKBAR", "2019-ag-6057": "KHADIJAH RASOOL", "2019-ag-6058": "ANSA SALEEM", "2019-ag-6059": "ZUHA QAMMAR", "2019-ag-6060": "SYED ASAD ALI BUKHARI", "2019-ag-6061": "MUHAMMAD USAMA SHABIR", "2019-ag-6062": "MUHAMMADSHAHZAIB", "2019-ag-6063": "BUSHRA HAMEED", "2019-ag-6065": "HINA FAROOQ", "2019-ag-6066": "ASIMA SHABBIR", "2019-ag-6067": "MUHAMMAD KHAWAR AZEEM", "2019-ag-6068": "MUHAMMAD AHTESHAM SARWAR", "2019-ag-6069": "MAIRA", "2019-ag-6070": "MAKARAM TAYYAB", "2019-ag-6071": "ABDULLAH ALTAF", "2019-ag-6072": "TALHA AZEEM", "2019-ag-6073": "MUHAMMAD NOMAN", "2019-ag-6074": "TASBIHA TANVEER", "2019-ag-6075": "SHAFA ZAMAN", "2019-ag-6076": "ZOHA USMAN", "2019-ag-6077": "MISHAL JAVAID", "2019-ag-6078": "NABEEL UR REHMAN", "2019-ag-6079": "MUHAMMAD ZAIN", "2019-ag-6080": "LAIBA", "2019-ag-6081": "MUHAMMAD UMER", "2019-ag-6082": "MOHAMMAD RUMAN WARIS", "2019-ag-6084": "SAMIA TANVEER", "2019-ag-6085": "MANAL KHALID", "2019-ag-6086": "HAFIZA KHADIJA SULEMAN", "2019-ag-6087": "MUHAMMAD HASSAN KHALID", "2019-ag-6088": "AASHIR NADEEM", "2019-ag-6089": "MUHAMMAD HAMZA IMRAN BAJWA", "2019-ag-6090": "ADEEL AHMAD", "2019-ag-6091": "MUHAMMAD ABUBAKAR", "2017-ag-7217": "RANA TALHA", "2017-ag-7297":"Ahmed Raza","2019-ag-6137":"Aneebb Hassan","2018-ag-8240":"Usama Shahbaz"}
dataFromFile = []
chatListDir = ''
fileName = ''
sg.theme("DarkBrown4")
layout =  [
          [sg.Text("Create New File if does not Already Exsist")],
          [sg.Button("Create New File")],
          [sg.Text("Choose the Chat File:   "), sg.Input(),sg.FileBrowse(key="-IN-",file_types=(("Text Files", "*.txt"),))],
          [sg.Text("Choose File to Update: "), sg.Input(),sg.FileBrowse(key="-OUT-",file_types=(("Excel Files", "*.xlsx"),))],
          [sg.Text('')],
          [sg.Button("Submit",size=(20,5)),sg.Button("Exit",size=(20,5))]
          ]

window = sg.Window('Attendance Checker', layout, size=(600, 200))

while True:
    event, values = window.read()
    if event == sg.WIN_CLOSED or event == "Exit":
        break
    elif event=="Create New File":
        new_file_dir = sg.popup_get_folder('Browse to folder to Create File')
        fileName=sg.popup_get_text('Enter the File Name')
        workbook=xlsxwriter.Workbook(new_file_dir+'/'+fileName+'.xlsx')
        worksheet=workbook.add_worksheet()
        worksheet.write(0,0,"Ag")
        worksheet.write(0,1,"Name")
        n = 1
        for i in agList:
             worksheet.write(n,0,i)
             worksheet.write(n,1,agList.get(i))
             n += 1
        workbook.close()
        sg.popup("File is Created at\n", new_file_dir)
            

    elif event == "Submit":
        chatListDir = values["-IN-"]
        outPutDir=values["-OUT-"]
        
        with open(chatListDir) as f:
            dataFromFile = np.loadtxt(f, dtype=str, delimiter="\n").tolist()


        wb=openpyxl.load_workbook(outPutDir)
        ws=wb.active
        cell_char='A'
        empty_cell_char=''
        while True:

            if (ws[str(cell_char)+"2"]).value==None:
                
                 empty_cell_char=str(cell_char)
                 break
            else:
                char_num=ord(cell_char[0])
                char_num+=1
                cell_char=chr(char_num)
        today=date.today()
        ws[empty_cell_char+str(1)]=today
        for i in range(2, len(agList)+2):
            for j in dataFromFile:
                agNum=(j[-12:])
                if(ws['A'+str(i)].value == agNum):
                    ws[empty_cell_char+str(i)] = "P"
                    break
                else:
                    ws[empty_cell_char+str(i)] = "A"

        wb.save(outPutDir)
        sg.popup('File Updated','New Data has been applied to File')
        break



